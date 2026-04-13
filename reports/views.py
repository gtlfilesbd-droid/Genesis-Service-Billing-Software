from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum
from billing.models import Bill
from billing.sync_auto_bills import sync_billing_queues
from clients.models import Client
from accounts.models import UserProfile
from datetime import date, timedelta, datetime
from django.utils import timezone
from decimal import Decimal
import calendar
import csv
import io
import json


def get_profile(user):
    try:
        return user.profile
    except UserProfile.DoesNotExist:
        return UserProfile.objects.create(user=user)


@login_required
def report_dashboard(request):
    profile = get_profile(request.user)
    if not (profile.can_view_reports or request.user.is_superuser or request.user.is_staff):
        from django.contrib import messages
        messages.error(request, 'You do not have permission to view reports.')
        from django.shortcuts import redirect
        return redirect('dashboard')

    sync_billing_queues()
    today = date.today()
    year_raw = (request.GET.get('year') or '').strip()
    year = int(year_raw) if year_raw.isdigit() else None
    month_raw = (request.GET.get('month') or '').strip()
    month = int(month_raw) if month_raw.isdigit() else None
    client_raw = (request.GET.get('client') or '').strip()
    client_id = int(client_raw) if client_raw.isdigit() else None
    status_filter = (request.GET.get('status') or '').strip().lower()
    if status_filter not in ('pending', 'submitted', 'paid'):
        status_filter = ''

    # Base queryset for filtered reporting (invoice_date-driven)
    filtered_qs = Bill.objects.select_related('client')
    # Scope queryset for KPIs/charts: apply client/year/month only (not status filter)
    scope_qs = Bill.objects.select_related('client')
    if client_id:
        filtered_qs = filtered_qs.filter(client_id=client_id)
        scope_qs = scope_qs.filter(client_id=client_id)
    if year is not None:
        filtered_qs = filtered_qs.filter(invoice_date__year=year)
        scope_qs = scope_qs.filter(invoice_date__year=year)
    if month:
        filtered_qs = filtered_qs.filter(invoice_date__month=month)
        scope_qs = scope_qs.filter(invoice_date__month=month)
    if status_filter:
        filtered_qs = filtered_qs.filter(status=status_filter)

    # Monthly revenue for chart — reflect client/year/month filters
    # - If year is selected: show that year's month breakdown (or single month if month selected)
    # - Else: show rolling last 12 months, filtered by client if selected
    chart_qs = Bill.objects.all()
    if client_id:
        chart_qs = chart_qs.filter(client_id=client_id)

    monthly_data = []
    if year is not None:
        months = [month] if month else list(range(1, 13))
        for m in months:
            label = date(year, m, 1).strftime('%b %Y')
            total = chart_qs.filter(
                status='paid',
                invoice_date__year=year,
                invoice_date__month=m,
            ).aggregate(t=Sum('total_in_bdt'))['t'] or 0
            monthly_data.append({'label': label, 'total': float(total)})
        chart_title = f'Monthly Revenue ({year})' if not month else f'Monthly Revenue ({year}-{m:02d})'
    else:
        for i in range(11, -1, -1):
            d = today.replace(day=1) - timedelta(days=i * 30)
            label = d.strftime('%b %Y')
            total = chart_qs.filter(
                status='paid',
                invoice_date__year=d.year,
                invoice_date__month=d.month,
            ).aggregate(t=Sum('total_in_bdt'))['t'] or 0
            monthly_data.append({'label': label, 'total': float(total)})
        chart_title = 'Monthly Revenue (Last 12 Months)'

    # Status breakdown (workflow) — reflect current scope filters (client/year/month)
    status_data = {
        'paid': scope_qs.filter(status='paid').count(),
        'pending': scope_qs.filter(status='pending').count(),
        'submitted': scope_qs.filter(status='submitted').count(),
    }

    # Top clients
    top_clients = Bill.objects.filter(status='paid').values(
        'client__name'
    ).annotate(total=Sum('total_in_bdt')).order_by('-total')[:5]

    # Summary
    total_revenue = Bill.objects.filter(status='paid').aggregate(t=Sum('total_in_bdt'))['t'] or 0
    pending_amount = Bill.objects.filter(
        status__in=['pending', 'submitted']
    ).aggregate(t=Sum('total_in_bdt'))['t'] or 0
    this_month_revenue = Bill.objects.filter(
        status='paid', invoice_date__year=today.year, invoice_date__month=today.month
    ).aggregate(t=Sum('total_in_bdt'))['t'] or 0

    # Filtered metrics (client/month/year/status)
    filtered_counts = {
        'pending': filtered_qs.filter(status='pending').count(),
        'submitted': filtered_qs.filter(status='submitted').count(),
        'paid': filtered_qs.filter(status='paid').count(),
        'total': filtered_qs.count(),
    }
    filtered_amounts = {
        'pending': filtered_qs.filter(status='pending').aggregate(t=Sum('total_in_bdt'))['t'] or 0,
        'submitted': filtered_qs.filter(status='submitted').aggregate(t=Sum('total_in_bdt'))['t'] or 0,
        'paid': filtered_qs.filter(status='paid').aggregate(t=Sum('total_in_bdt'))['t'] or 0,
        'total': filtered_qs.aggregate(t=Sum('total_in_bdt'))['t'] or 0,
    }

    # Expected vs received vs due within the current scope (client/year/month)
    expected_scope = scope_qs.aggregate(t=Sum('total_in_bdt'))['t'] or 0
    received_scope = scope_qs.filter(status='paid').aggregate(t=Sum('total_in_bdt'))['t'] or 0
    due_scope = scope_qs.filter(status__in=['pending', 'submitted']).aggregate(t=Sum('total_in_bdt'))['t'] or 0

    if year is not None and month:
        scope_label = f'{year}-{month:02d}'
    elif year is not None:
        scope_label = str(year)
    elif month:
        scope_label = f'All years ({calendar.month_name[month]})'
    else:
        scope_label = 'All years'

    recent_bills = filtered_qs.order_by('-invoice_date', '-id')[:200]

    # Year options for filter dropdown
    year_options = list(
        Bill.objects.order_by()
        .values_list('invoice_date__year', flat=True)
        .distinct()
    )
    year_options = [y for y in year_options if y is not None]
    year_options.sort(reverse=True)

    filters_active = bool(client_id or year is not None or month or status_filter)

    # Filter-aware KPI cards (top section on the page)
    scope_total_revenue = scope_qs.filter(status='paid').aggregate(t=Sum('total_in_bdt'))['t'] or 0
    scope_outstanding = scope_qs.filter(status__in=['pending', 'submitted']).aggregate(t=Sum('total_in_bdt'))['t'] or 0
    scope_total_bills = scope_qs.count()
    scope_client_count = scope_qs.values('client_id').distinct().count()

    # "This Month" KPI becomes "This period" when month/year filters are used
    if year is None and not month:
        period_year = today.year
        period_month = today.month
    else:
        period_year = year if year is not None else today.year
        period_month = month if month else today.month
    scope_period_revenue = scope_qs.filter(
        status='paid',
        invoice_date__year=period_year,
        invoice_date__month=period_month,
    ).aggregate(t=Sum('total_in_bdt'))['t'] or 0

    context = {
        'profile': profile,
        'monthly_data_json': json.dumps(monthly_data),
        'monthly_chart_title': chart_title,
        'status_data_json': json.dumps(status_data),
        'top_clients': top_clients,
        'total_revenue': total_revenue,
        'pending_amount': pending_amount,
        'this_month_revenue': this_month_revenue,
        'total_clients': Client.objects.filter(is_active=True).count(),
        'total_bills': Bill.objects.count(),
        'year': year,
        'month': month,
        'client_id': client_id,
        'status': status_filter,
        'clients': Client.objects.filter(is_active=True).order_by('name'),
        'year_options': year_options,
        'filtered_counts': filtered_counts,
        'filtered_amounts': filtered_amounts,
        'expected_scope': expected_scope,
        'received_scope': received_scope,
        'due_scope': due_scope,
        'expected_scope_label': scope_label,
        'recent_bills': recent_bills,
        'filters_active': filters_active,
        'scope_total_revenue': scope_total_revenue,
        'scope_outstanding': scope_outstanding,
        'scope_period_revenue': scope_period_revenue,
        'scope_total_bills': scope_total_bills,
        'scope_client_count': scope_client_count,
        'scope_period_year': period_year,
        'scope_period_month': period_month,
    }
    return render(request, 'reports/report_dashboard.html', context)


def _export_bills_queryset(request):
    """Bills for export, same filters for CSV and Excel."""
    qs = Bill.objects.select_related('client').order_by('-invoice_date', '-id')
    status_filter = (request.GET.get('status') or '').strip()
    if status_filter in ('paid', 'pending', 'submitted'):
        qs = qs.filter(status=status_filter)
    client_raw = (request.GET.get('client') or '').strip()
    if client_raw.isdigit():
        qs = qs.filter(client_id=int(client_raw))
    year_raw = (request.GET.get('year') or '').strip()
    if year_raw.isdigit():
        qs = qs.filter(invoice_date__year=int(year_raw))
    month_raw = (request.GET.get('month') or '').strip()
    if month_raw.isdigit():
        qs = qs.filter(invoice_date__month=int(month_raw))
    return qs


def _bill_export_tuple(bill):
    """One row: values aligned with BILL_EXPORT_HEADERS (Excel-friendly types)."""
    c = bill.client
    company = (c.company or '').strip() if getattr(c, 'company', None) else ''

    def dval(d):
        return d if d is not None else ''

    def fval(x):
        if x is None:
            return Decimal('0')
        return x

    return (
        bill.bill_number or '',
        (bill.invoice_number or '').strip(),
        c.name or '',
        company,
        dval(bill.invoice_date),
        dval(bill.po_date),
        dval(bill.bill_period_from),
        dval(bill.bill_period_to),
        (bill.bill_period or '').strip(),
        (bill.service_period or '').strip(),
        bill.get_status_display(),
        fval(bill.subtotal),
        fval(bill.project_base_value),
        fval(bill.vat_rate_percent),
        fval(bill.vat_amount),
        fval(bill.ait_rate_percent),
        fval(bill.ait_amount),
        fval(bill.excluding_vat_ait),
        fval(bill.total_in_bdt),
        dval(bill.payment_date),
        (bill.payment_method or '').strip(),
        (bill.payment_reference or '').strip(),
        'Yes' if bill.auto_generated else 'No',
        bill.created_at,
    )


BILL_EXPORT_HEADERS = (
    'Bill reference',
    'Invoice number',
    'Client name',
    'Client company',
    'Invoice date',
    'PO / due date',
    'Bill period from',
    'Bill period to',
    'Bill period (description)',
    'Service period',
    'Status',
    'Items subtotal (BDT)',
    'Base value (BDT)',
    'VAT rate (%)',
    'VAT amount (BDT)',
    'AIT rate (%)',
    'AIT amount (BDT)',
    'Total VAT + AIT (BDT)',
    'Grand total (BDT)',
    'Payment date',
    'Payment method',
    'Payment reference',
    'Auto-generated',
    'Created at',
)


def _csv_cell(v):
    if v is None or v == '':
        return ''
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, Decimal):
        return format(v, 'f')
    return str(v)


def _excel_naive_datetime(val):
    """openpyxl rejects tz-aware datetimes; store as local wall time without tzinfo."""
    if val is None:
        return None
    if isinstance(val, datetime):
        if timezone.is_aware(val):
            return timezone.localtime(val).replace(tzinfo=None)
        return val
    return val


@login_required
def export_bills_csv(request):
    profile = get_profile(request.user)
    if not (profile.can_export_reports or request.user.is_superuser or request.user.is_staff):
        from django.contrib import messages
        from django.shortcuts import redirect
        messages.error(request, 'Permission denied.')
        return redirect('report_dashboard')

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(BILL_EXPORT_HEADERS)

    for bill in _export_bills_queryset(request):
        row = _bill_export_tuple(bill)
        writer.writerow([_csv_cell(x) for x in row])

    payload = '\ufeff' + buf.getvalue()
    response = HttpResponse(payload, content_type='text/csv; charset=utf-8')
    st = (request.GET.get('status') or '').strip()
    suffix = f'-{st}' if st in ('paid', 'pending', 'submitted') else ''
    response['Content-Disposition'] = (
        f'attachment; filename="bills-export{suffix}-{date.today().isoformat()}.csv"'
    )
    return response


@login_required
def export_bills_excel(request):
    profile = get_profile(request.user)
    if not (profile.can_export_reports or request.user.is_superuser or request.user.is_staff):
        from django.contrib import messages
        from django.shortcuts import redirect
        messages.error(request, 'Permission denied.')
        return redirect('report_dashboard')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Bills'

        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hfill = PatternFill('solid', fgColor='1565C0')
        hfont = Font(bold=True, color='FFFFFF', size=11)
        meta_font = Font(size=10, color='444444')

        st = (request.GET.get('status') or '').strip()
        filter_note = {
            'paid': 'Paid bills only',
            'pending': 'Pending bills only',
            'submitted': 'Submitted bills only',
        }.get(st, 'All statuses')

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(BILL_EXPORT_HEADERS))
        t1 = ws.cell(row=1, column=1, value='Genesis BillSoft — Bills export')
        t1.font = Font(bold=True, size=14, color='1565C0')
        t1.alignment = Alignment(horizontal='left', vertical='center')

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(BILL_EXPORT_HEADERS))
        gen = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M %Z')
        t2 = ws.cell(row=2, column=1, value=f'Generated: {gen}  |  Filter: {filter_note}')
        t2.font = meta_font
        t2.alignment = Alignment(horizontal='left', vertical='center')

        header_row = 4
        for col, title in enumerate(BILL_EXPORT_HEADERS, start=1):
            cell = ws.cell(row=header_row, column=col, value=title)
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        money_cols = {12, 13, 15, 17, 18, 19}  # 1-based column index in sheet
        pct_cols = {14, 16}
        date_cols = {5, 6, 7, 8, 20}
        dt_cols = {24}

        data_row = header_row + 1
        for bill in _export_bills_queryset(request):
            row_vals = _bill_export_tuple(bill)
            for col, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=data_row, column=col)
                cell.border = border
                if col in date_cols:
                    cell.value = val if val else None
                    cell.number_format = 'YYYY-MM-DD'
                elif col in dt_cols:
                    cell.value = _excel_naive_datetime(val)
                    cell.number_format = 'YYYY-MM-DD HH:MM'
                elif col in money_cols:
                    cell.value = float(val) if val is not None else 0.0
                    cell.number_format = '#,##0.00'
                elif col in pct_cols:
                    cell.value = float(val) if val is not None else 0.0
                    cell.number_format = '0.00'
                else:
                    cell.value = val
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            data_row += 1

        widths = (
            14, 18, 22, 20, 12, 12, 12, 12, 28, 20, 11,
            14, 14, 10, 12, 10, 12, 14, 14, 12, 14, 14, 12, 18,
        )
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[header_row].height = 36

        suffix = f'-{st}' if st in ('paid', 'pending', 'submitted') else ''
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = (
            f'attachment; filename="bills-export{suffix}-{date.today().isoformat()}.xlsx"'
        )
        wb.save(response)
        return response
    except ImportError:
        from django.contrib import messages
        from django.shortcuts import redirect
        messages.error(request, 'openpyxl not installed.')
        return redirect('report_dashboard')
