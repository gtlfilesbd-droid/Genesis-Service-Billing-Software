from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.utils import timezone
from django.db import transaction
from django.db.models import Prefetch, Q

from .models import Bill, BillItem, BillingBank, BillingTaxSettings
from .invoice_number import build_invoice_number_base
from .bill_period import compute_bill_period_window, format_bill_period_line
from clients.models import Client, Agreement, Service
from accounts.models import UserProfile, AuditLog
import calendar
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation

from django.utils.dateparse import parse_date


def _safe_excel_sheet_title(raw, prefix='Inv', max_len=31):
    """openpyxl rejects sheet names containing : \\ / ? * [ ] and limits length to 31."""
    s = str(raw or '').strip() or 'inv'
    for c in r':\/?*[]':
        s = s.replace(c, '-')
    s = ' '.join(s.split())
    title = f'{prefix}-{s}'.strip('-')[:max_len]
    return title if title else 'Invoice'


def _safe_download_filename(raw, max_len=150):
    """Remove characters unsafe in Windows / browser download filenames."""
    s = str(raw or 'invoice').strip() or 'invoice'
    for c in r'<>:"/\|?*[]':
        s = s.replace(c, '-')
    s = ' '.join(s.split()).strip('-') or 'invoice'
    return s[:max_len]


def _parse_rate_percent(request, key, default):
    raw = request.POST.get(key)
    if raw is None or str(raw).strip() == '':
        return default
    try:
        v = Decimal(str(raw).strip().replace(',', '.'))
        if v < 0:
            v = Decimal('0')
        if v > Decimal('999.99'):
            v = Decimal('999.99')
        return v
    except (InvalidOperation, ValueError, TypeError):
        return default


def _bill_form_extra_context(bill=None):
    s = BillingTaxSettings.get_solo()
    banks = list(BillingBank.objects.all().order_by('-is_default', 'label'))
    payload = [
        {
            'id': b.pk,
            'label': b.label,
            'bank_name': b.bank_name or '',
            'beneficiary': b.beneficiary or '',
            'bank_branch': b.bank_branch or '',
            'bank_address_line1': b.bank_address_line1 or '',
            'bank_address_line2': b.bank_address_line2 or '',
            'account_number': b.account_number or '',
            'swift_code': b.swift_code or '',
            'branch_routing_code': b.branch_routing_code or '',
            'bin_number': b.bin_number or '',
            'tin_number': b.tin_number or '',
        }
        for b in banks
    ]
    initial = None
    if bill and getattr(bill, 'billing_bank_id', None):
        initial = bill.billing_bank_id
    else:
        for b in banks:
            if b.is_default:
                initial = b.pk
                break
    return {
        'tax_vat_percent': float(s.vat_percent),
        'tax_ait_percent': float(s.ait_percent),
        'billing_banks': banks,
        'billing_banks_payload': payload,
        'initial_billing_bank_id': initial,
    }


def _set_billing_bank_fk_from_post(request, bill):
    raw = request.POST.get('billing_bank')
    bb = None
    if raw is not None and str(raw).strip() != '':
        bb = BillingBank.objects.filter(pk=raw).first()
    elif raw is not None and str(raw).strip() == '':
        bb = None
    else:
        bb = BillingBank.get_default()
    bill.billing_bank = bb


def _apply_bank_fields_from_post(request, bill):
    """Bank text fields come from the bill form (editable per invoice)."""
    bill.bank_name = (request.POST.get('bank_name') or '').strip()
    bill.beneficiary = (request.POST.get('beneficiary') or '').strip()
    bill.bank_branch = (request.POST.get('bank_branch') or '').strip()
    bill.bank_address_line1 = (request.POST.get('bank_address_line1') or '').strip()
    bill.bank_address_line2 = (request.POST.get('bank_address_line2') or '').strip()
    bill.account_number = (request.POST.get('account_number') or '').strip()
    bill.swift_code = (request.POST.get('swift_code') or '').strip()
    bill.branch_routing_code = (request.POST.get('branch_routing_code') or '').strip()
    bill.bin_number = (request.POST.get('bin_number') or '').strip()
    bill.tin_number = (request.POST.get('tin_number') or '').strip()


def get_profile(user):
    try:
        return user.profile
    except UserProfile.DoesNotExist:
        return UserProfile.objects.create(user=user)


def _render_queue_grouped(request, bills_base_qs, template_name, profile):
    from .sync_auto_bills import sync_billing_queues, group_bills_by_invoice_month

    sync_billing_queues()
    status_filter = request.GET.get('status', '')
    client_filter = request.GET.get('client', '')
    search = request.GET.get('search', '')
    inv_year = request.GET.get('inv_year', '').strip()
    inv_month = request.GET.get('inv_month', '').strip()

    bills = bills_base_qs.select_related('client')
    if status_filter:
        bills = bills.filter(status=status_filter)
    if client_filter:
        bills = bills.filter(client_id=client_filter)
    if search:
        bills = bills.filter(
            Q(bill_number__icontains=search)
            | Q(client__name__icontains=search)
            | Q(invoice_number__icontains=search)
        )
    if inv_year.isdigit():
        bills = bills.filter(invoice_date__year=int(inv_year))
    if inv_month.isdigit():
        m = int(inv_month)
        if 1 <= m <= 12:
            bills = bills.filter(invoice_date__month=m)

    grouped = group_bills_by_invoice_month(bills)

    year_rows = Bill.objects.exclude(invoice_date__isnull=True).dates(
        'invoice_date', 'year', order='DESC'
    )
    invoice_year_choices = [d.year for d in year_rows]
    if not invoice_year_choices:
        invoice_year_choices = [date.today().year]
    month_choices = [(str(i), calendar.month_name[i]) for i in range(1, 13)]

    return render(
        request,
        template_name,
        {
            'grouped_bills': grouped,
            'clients': Client.objects.filter(is_active=True),
            'status_filter': status_filter,
            'client_filter': client_filter,
            'search': search,
            'inv_year': inv_year,
            'inv_month': inv_month,
            'invoice_year_choices': invoice_year_choices,
            'month_choices': month_choices,
            'status_choices': Bill._meta.get_field('status').choices,
            'profile': profile,
        },
    )


def _validate_bill_invoice_prerequisites(client_id, agreement_id):
    if not client_id or not agreement_id:
        return False, 'Please select a client and an agreement.'
    ag = Agreement.objects.select_related('agreement_with').filter(pk=agreement_id).first()
    if not ag:
        return False, 'Invalid agreement.'
    if not ag.agreement_with_id:
        return False, 'Selected agreement must have an Agreement With company.'
    cl = Client.objects.filter(pk=client_id).first()
    if not cl:
        return False, 'Invalid client.'
    if not (cl.short_form or '').strip():
        return False, 'Client must have a short form before creating a bill.'
    return True, None


def _save_bill_from_post(request, bill):
    with transaction.atomic():
        tax_defaults = BillingTaxSettings.get_solo()
        bill.vat_rate_percent = _parse_rate_percent(
            request, 'vat_rate_percent', tax_defaults.vat_percent
        )
        bill.ait_rate_percent = _parse_rate_percent(
            request, 'ait_rate_percent', tax_defaults.ait_percent
        )
        bill.client_id = request.POST.get('client')
        bill.agreement_id = request.POST.get('agreement') or None
        inv_raw = request.POST.get('invoice_date')
        if inv_raw:
            bill.invoice_date = parse_date(inv_raw) or date.today()
        else:
            bill.invoice_date = date.today()
        inv = bill.invoice_date
        is_edit = request.POST.get('bill_form_is_edit') == '1'

        if is_edit:
            bill._skip_auto_invoice_number = True
            inv_num = (request.POST.get('invoice_number') or '').strip()
            if inv_num:
                bill.invoice_number = inv_num
            po_raw = request.POST.get('po_date')
            bill.po_date = parse_date(po_raw) if po_raw else None
            bf_raw = request.POST.get('bill_period_from')
            bt_raw = request.POST.get('bill_period_to')
            bill.bill_period_from = parse_date(bf_raw) if bf_raw else None
            bill.bill_period_to = parse_date(bt_raw) if bt_raw else None
            bill.bill_period = format_bill_period_line(bill.bill_period_from, bill.bill_period_to)
        elif bill.agreement_id:
            try:
                ag = Agreement.objects.prefetch_related(
                    Prefetch('services', queryset=Service.objects.filter(is_active=True).order_by('id'))
                ).get(pk=bill.agreement_id)
                bill.po_date = ag.start_date
                bf, bt = compute_bill_period_window(ag, inv)
                bill.bill_period_from = bf
                bill.bill_period_to = bt
                bill.bill_period = format_bill_period_line(bf, bt)
            except Agreement.DoesNotExist:
                bill.po_date = None
                bill.bill_period_from = None
                bill.bill_period_to = None
                bill.bill_period = ''
        else:
            bill.po_date = None
            bill.bill_period_from = None
            bill.bill_period_to = None
            bill.bill_period = ''
        bill.service_period = ''
        if 'project_value_yearly' in request.POST:
            bill.project_value_yearly = request.POST.get('project_value_yearly') or 0
        bill.remark = request.POST.get('remark', '')
        _set_billing_bank_fk_from_post(request, bill)
        _apply_bank_fields_from_post(request, bill)
        if not is_edit:
            bill.status = 'pending'
        if not bill.created_by_id:
            bill.created_by = request.user
        bill.save()

        bill.items.all().delete()
        descriptions = request.POST.getlist('item_description[]')
        quantities = request.POST.getlist('item_quantity[]')
        units = request.POST.getlist('item_unit[]')
        prices = request.POST.getlist('item_price[]')

        for i in range(len(descriptions)):
            desc = descriptions[i].strip() if i < len(descriptions) else ""
            if not desc:
                continue
            try:
                qty = float(quantities[i]) if i < len(quantities) and quantities[i] else 1
            except (ValueError, TypeError):
                qty = 1
            unit = units[i] if i < len(units) else ""
            try:
                price = float(prices[i]) if i < len(prices) and prices[i] else 0
            except (ValueError, TypeError):
                price = 0
            BillItem.objects.create(
                bill=bill,
                description=desc,
                quantity=qty,
                unit=unit,
                unit_price=price,
            )
        bill.calculate_totals()
        bill.save(
            update_fields=[
                'subtotal',
                'project_base_value',
                'vat_rate_percent',
                'ait_rate_percent',
                'vat_amount',
                'ait_amount',
                'excluding_vat_ait',
                'total_in_bdt',
            ]
        )
    return bill


@login_required
def bill_list(request):
    from .sync_auto_bills import sync_billing_queues

    sync_billing_queues()
    bills = Bill.objects.select_related('client').all()
    status_filter = request.GET.get('status', '')
    client_filter = request.GET.get('client', '')
    search = request.GET.get('search', '')
    inv_year = request.GET.get('inv_year', '').strip()
    inv_month = request.GET.get('inv_month', '').strip()
    if status_filter:
        bills = bills.filter(status=status_filter)
    if client_filter:
        bills = bills.filter(client_id=client_filter)
    if search:
        bills = bills.filter(
            Q(bill_number__icontains=search)
            | Q(client__name__icontains=search)
            | Q(invoice_number__icontains=search)
        )
    if inv_year.isdigit():
        bills = bills.filter(invoice_date__year=int(inv_year))
    if inv_month.isdigit():
        m = int(inv_month)
        if 1 <= m <= 12:
            bills = bills.filter(invoice_date__month=m)

    filtered_total = bills.count()
    list_stats = {
        'total': filtered_total,
        'pending': bills.filter(status='pending').count(),
        'submitted': bills.filter(status='submitted').count(),
        'paid': bills.filter(status='paid').count(),
    }

    paginator = Paginator(bills.order_by('-invoice_date', '-id'), 15)
    bills_page = paginator.get_page(request.GET.get('page'))
    profile = get_profile(request.user)

    q = request.GET.copy()
    q.pop('page', None)
    filter_query = q.urlencode()

    year_rows = Bill.objects.exclude(invoice_date__isnull=True).dates(
        'invoice_date', 'year', order='DESC'
    )
    invoice_year_choices = [d.year for d in year_rows]
    if not invoice_year_choices:
        invoice_year_choices = [date.today().year]
    month_choices = [(str(i), calendar.month_name[i]) for i in range(1, 13)]

    return render(request, 'bills/bill_list.html', {
        'bills': bills_page,
        'clients': Client.objects.filter(is_active=True),
        'status_filter': status_filter,
        'client_filter': client_filter,
        'search': search,
        'inv_year': inv_year,
        'inv_month': inv_month,
        'invoice_year_choices': invoice_year_choices,
        'month_choices': month_choices,
        'profile': profile,
        'status_choices': Bill._meta.get_field('status').choices,
        'list_stats': list_stats,
        'filter_query': filter_query,
    })


@login_required
def bill_create(request):
    profile = get_profile(request.user)
    if not profile.can_generate_bill and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to generate bills.')
        return redirect('bill_list')
    clients = Client.objects.filter(is_active=True)
    if request.method == 'POST':
        if not request.POST.get('client'):
            messages.error(request, 'Please select a client.')
        elif not request.POST.get('agreement'):
            messages.error(request, 'Please select an agreement.')
        else:
            ok, err = _validate_bill_invoice_prerequisites(
                request.POST.get('client'), request.POST.get('agreement')
            )
            if not ok:
                messages.error(request, err)
            else:
                bill = Bill()
                _save_bill_from_post(request, bill)
                inv = bill.invoice_number or bill.bill_number
                messages.success(request, f'Invoice {inv} created successfully.')
                return redirect('bill_detail', pk=bill.pk)
    ctx = {
        'clients': clients, 'profile': profile,
        'today': date.today(),
        'status_choices': Bill._meta.get_field('status').choices,
        'action': 'Create',
    }
    ctx.update(_bill_form_extra_context())
    return render(request, 'bills/bill_form.html', ctx)


@login_required
def bill_detail(request, pk):
    from .sync_auto_bills import sync_billing_queues

    sync_billing_queues()
    bill = get_object_or_404(
        Bill.objects.select_related('client', 'agreement', 'agreement__agreement_with'),
        pk=pk,
    )
    profile = get_profile(request.user)
    return render(request, 'bills/bill_detail.html', {'bill': bill, 'profile': profile})


@login_required
def bill_edit(request, pk):
    profile = get_profile(request.user)
    bill = get_object_or_404(Bill, pk=pk)
    if not profile.can_edit_bill and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to edit bills.')
        return redirect('bill_detail', pk=pk)
    if bill.status in ('submitted', 'paid'):
        messages.error(
            request,
            'This invoice cannot be edited after it has been submitted to the client or marked paid.',
        )
        return redirect('bill_detail', pk=pk)
    clients = Client.objects.filter(is_active=True)
    if request.method == 'POST':
        if not request.POST.get('agreement'):
            messages.error(request, 'Please select an agreement.')
            return redirect('bill_edit', pk=pk)
        ok, err = _validate_bill_invoice_prerequisites(
            request.POST.get('client'), request.POST.get('agreement')
        )
        if not ok:
            messages.error(request, err)
            ctx = {
                'bill': bill, 'clients': clients, 'profile': profile,
                'today': date.today(),
                'status_choices': Bill._meta.get_field('status').choices,
                'action': 'Edit',
            }
            ctx.update(_bill_form_extra_context(bill))
            return render(request, 'bills/bill_form.html', ctx)
        _save_bill_from_post(request, bill)
        inv = bill.invoice_number or bill.bill_number
        messages.success(request, f'Invoice {inv} updated successfully.')
        return redirect('bill_detail', pk=pk)
    ctx = {
        'bill': bill, 'clients': clients, 'profile': profile,
        'today': date.today(),
        'status_choices': Bill._meta.get_field('status').choices,
        'action': 'Edit',
    }
    ctx.update(_bill_form_extra_context(bill))
    return render(request, 'bills/bill_form.html', ctx)


@login_required
def get_client_agreements(request, client_id):
    client = get_object_or_404(Client, pk=client_id)
    agreements = client.agreements.filter(is_active=True).prefetch_related(
        Prefetch('services', queryset=Service.objects.filter(is_active=True).order_by('id'))
    ).order_by('-start_date', '-id')
    payload = [{
        'id': a.id,
        'title': a.title,
        'start_date': a.start_date.isoformat() if a.start_date else '',
        'end_date': a.end_date.isoformat() if a.end_date else '',
    } for a in agreements]
    return JsonResponse({'agreements': payload})


@login_required
def preview_bill_period(request):
    agreement_id = request.GET.get('agreement')
    inv_raw = request.GET.get('invoice_date')
    if not agreement_id:
        return JsonResponse({'bill_period_from': '', 'bill_period_to': '', 'summary': ''})
    inv = parse_date(inv_raw) if inv_raw else date.today()
    if inv is None:
        inv = date.today()
    ag = get_object_or_404(
        Agreement.objects.prefetch_related(
            Prefetch('services', queryset=Service.objects.filter(is_active=True).order_by('id'))
        ),
        pk=agreement_id,
    )
    bf, bt = compute_bill_period_window(ag, inv)
    return JsonResponse({
        'bill_period_from': bf.isoformat() if bf else '',
        'bill_period_to': bt.isoformat() if bt else '',
        'summary': format_bill_period_line(bf, bt),
    })


@login_required
def preview_invoice_number(request):
    client_id = request.GET.get('client')
    agreement_id = request.GET.get('agreement')
    invoice_date = request.GET.get('invoice_date')
    if not client_id or not agreement_id or not invoice_date:
        return JsonResponse({'preview': '', 'error': None})
    ok, err = _validate_bill_invoice_prerequisites(client_id, agreement_id)
    if not ok:
        return JsonResponse({'preview': '', 'error': err})
    try:
        ag = Agreement.objects.select_related('agreement_with').get(pk=agreement_id)
        client = Client.objects.get(pk=client_id)
        preview = build_invoice_number_base(ag, client, invoice_date)
    except (Agreement.DoesNotExist, Client.DoesNotExist, ValueError) as e:
        return JsonResponse({'preview': '', 'error': str(e) or 'Could not build invoice number.'})
    return JsonResponse({'preview': preview, 'error': None})


@login_required
def get_agreement_services(request, agreement_id):
    agreement = get_object_or_404(Agreement, pk=agreement_id)
    services = [{'id': s.id, 'name': s.name, 'service_type': s.get_service_type_display(),
                 'charge': str(s.charge), 'description': s.description or ''}
                for s in agreement.services.filter(is_active=True)]
    return JsonResponse({'services': services, 'agreement_title': agreement.title})


def _bill_agreement_meta(bill):
    ag = getattr(bill, 'agreement', None)
    if not ag:
        return {'agreement_title': '', 'agreement_service_types': ''}
    types = []
    try:
        for s in ag.services.filter(is_active=True):
            t = (s.get_service_type_display() or '').strip()
            if t and t not in types:
                types.append(t)
    except Exception:
        types = []
    return {
        'agreement_title': (ag.title or '').strip(),
        'agreement_service_types': ', '.join(types),
    }


@login_required
def bill_pdf(request, pk):
    from io import BytesIO

    from django.template.loader import render_to_string

    bill = get_object_or_404(Bill, pk=pk)
    base_url = request.build_absolute_uri('/')
    ctx = {'bill': bill, 'letterhead_print': True}
    ctx.update(_bill_agreement_meta(bill))
    html_string = render_to_string('bills/bill_pdf.html', ctx)

    pdf_bytes = None
    try:
        import weasyprint

        pdf_bytes = weasyprint.HTML(string=html_string, base_url=base_url).write_pdf()
    except (ImportError, OSError):
        pass

    if pdf_bytes is None:
        try:
            from xhtml2pdf import pisa

            out = BytesIO()
            doc = pisa.CreatePDF(html_string, dest=out, encoding='utf-8')
            if not doc.err:
                pdf_bytes = out.getvalue()
        except ImportError:
            pass

    if pdf_bytes is None:
        messages.error(
            request,
            'PDF engine not available. Run: pip install xhtml2pdf '
            '(or install WeasyPrint with GTK on Windows — see WeasyPrint docs).',
        )
        return redirect('bill_detail', pk=pk)

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    fname = (bill.invoice_number or bill.bill_number).replace('/', '-')
    response['Content-Disposition'] = f'attachment; filename="Invoice-{fname}.pdf"'
    return response


@login_required
def bill_print(request, pk):
    bill = get_object_or_404(Bill, pk=pk)
    ctx = {'bill': bill, 'print_mode': True, 'letterhead_print': True}
    ctx.update(_bill_agreement_meta(bill))
    return render(
        request,
        'bills/bill_pdf.html',
        ctx,
    )


@login_required
def bill_excel(request, pk):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        from billing.money_words import bdt_amount_in_words

        bill = get_object_or_404(Bill, pk=pk)
        inv_disp = bill.invoice_number or bill.bill_number
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _safe_excel_sheet_title(inv_disp)
        thin = Side(style='thin', color='000000')
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
        bf = Font(bold=True)
        bfh = Font(bold=True, size=10)

        def apply_border_row(row_idx, cols=6):
            for col in range(1, cols + 1):
                ws.cell(row=row_idx, column=col).border = border_all

        ws.merge_cells('A1:F1')
        ws['A1'] = 'INVOICE'
        ws['A1'].font = Font(size=16, bold=True, color='000000')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:F2')
        ws['A2'] = f'Invoice #: {inv_disp}'
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.append([])
        ws.append(['Bill To:', '', '', 'Invoice Details:', '', ''])
        ws['A4'].font = bf
        ws['D4'].font = bf
        ws.append([bill.client.name, '', '', 'Invoice Date:', str(bill.invoice_date), ''])
        ws.append([bill.client.address, '', '', 'PO Date:', str(bill.po_date or ''), ''])
        if bill.bill_period_from or bill.bill_period_to:
            ws.append([bill.client.phone or '', '', '', 'Bill From:', str(bill.bill_period_from or ''), ''])
            ws.append([bill.client.email or '', '', '', 'Bill To:', str(bill.bill_period_to or ''), ''])
        else:
            ws.append([bill.client.phone or '', '', '', 'Bill Period:', bill.bill_period or '', ''])
            ws.append([bill.client.email or '', '', '', '', '', ''])
        ws.append([])
        headers = ['#', 'Description', 'Qty', 'Unit', 'Unit Price (BDT)', 'Amount (BDT)']
        ws.append(headers)
        hr = ws.max_row
        for col in range(1, 7):
            c = ws.cell(row=hr, column=col)
            c.font = bfh
            c.alignment = Alignment(horizontal='center', wrap_text=True)
            c.border = border_all
        first_item_row = hr + 1
        for idx, item in enumerate(bill.items.all(), 1):
            ws.append([idx, item.description, float(item.quantity), item.unit or '', float(item.unit_price), float(item.amount)])
        last_item_row = ws.max_row
        for r in range(first_item_row, last_item_row + 1):
            apply_border_row(r)
        ws.append([])
        ws.append(['', '', '', '', 'Items Subtotal:', float(bill.subtotal)])
        sub_row = ws.max_row
        apply_border_row(sub_row)
        ws.cell(row=sub_row, column=5).font = bf
        ws.cell(row=sub_row, column=6).font = bf
        ws.append([])
        ws.append(['VAT & AIT Details'])
        ws[f'A{ws.max_row}'].font = bf
        ws.append(['Base Value (BDT):', '', '', '', '', float(bill.project_base_value)])
        apply_border_row(ws.max_row)
        ws.append([f'VAT ({bill.vat_rate_percent}%):', '', '', '', '', float(bill.vat_amount)])
        apply_border_row(ws.max_row)
        ws.append([f'AIT ({bill.ait_rate_percent}%):', '', '', '', '', float(bill.ait_amount)])
        apply_border_row(ws.max_row)
        ws.append(['Total VAT & AIT:', '', '', '', '', float(bill.excluding_vat_ait)])
        apply_border_row(ws.max_row)
        ws.append(['Total In BDT (Base + VAT + AIT):', '', '', '', '', float(bill.total_in_bdt)])
        tot_row = ws.max_row
        for col in range(1, 7):
            c = ws.cell(row=tot_row, column=col)
            c.border = border_all
            c.font = bf
        ws.append(['In words:', bdt_amount_in_words(bill.total_in_bdt), '', '', '', ''])
        words_row = ws.max_row
        ws.merge_cells(start_row=words_row, start_column=2, end_row=words_row, end_column=6)
        apply_border_row(words_row)
        ws.cell(row=words_row, column=1).font = bf
        ws.cell(row=words_row, column=2).alignment = Alignment(wrap_text=True, vertical='top')
        if bill.bank_name:
            ws.append([])
            ws.append(['Bank information (For Payment)'])
            ws[f'A{ws.max_row}'].font = bf
            ws.append(['Bank Name:', bill.bank_name])
            ws.append(['Beneficiary:', bill.beneficiary or ''])
            ws.append(['Branch:', bill.bank_branch or ''])
            _addr_bits = [
                (bill.bank_address_line1 or '').strip(),
                (bill.bank_address_line2 or '').strip(),
            ]
            _addr = ', '.join(x for x in _addr_bits if x)
            if _addr:
                ws.append(['Address:', _addr])
            ws.append(['Account No:', bill.account_number or ''])
            ws.append(['Swift Code:', bill.swift_code or ''])
            ws.append(['Branch Code (Routing):', bill.branch_routing_code or ''])
            ws.append(['BIN:', bill.bin_number or ''])
            ws.append(['TIN:', bill.tin_number or ''])
        for col, w in zip('ABCDEF', [28, 40, 10, 14, 22, 22]):
            ws.column_dimensions[col].width = w
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        fname = _safe_download_filename(bill.invoice_number or bill.bill_number)
        response['Content-Disposition'] = f'attachment; filename="Invoice-{fname}.xlsx"'
        wb.save(response)
        return response
    except ImportError:
        messages.error(request, 'openpyxl not installed.')
        return redirect('bill_detail', pk=pk)


@login_required
def mark_paid(request, pk):
    if request.method != 'POST':
        messages.error(request, 'Use the Mark paid button on the bill page.')
        return redirect('bill_detail', pk=pk)
    profile = get_profile(request.user)
    if not profile.can_mark_bill_paid and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to mark bills as paid.')
        return redirect('bill_detail', pk=pk)
    bill = get_object_or_404(Bill, pk=pk)
    if bill.status != 'submitted':
        messages.error(request, 'Only submitted bills can be marked paid.')
        return redirect('bill_detail', pk=pk)
    bill.status = 'paid'
    bill.payment_date = date.today()
    bill.save()
    inv = bill.invoice_number or bill.bill_number
    messages.success(request, f'Invoice {inv} marked as paid.')
    return redirect('bill_detail', pk=pk)


@login_required
def bill_submit(request, pk):
    if request.method != 'POST':
        return redirect('bill_detail', pk=pk)
    profile = get_profile(request.user)
    if not profile.can_submit_bill and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to submit bills to the client.')
        return redirect('bill_detail', pk=pk)
    bill = get_object_or_404(Bill, pk=pk)
    if bill.status != 'pending':
        messages.error(request, 'Only pending bills can be submitted to the client.')
        return redirect('bill_detail', pk=pk)
    if not bill.is_mature:
        messages.error(
            request,
            'This bill is not mature yet (service period has not ended). You can submit after Bill To.',
        )
        return redirect('bill_detail', pk=pk)
    bill.status = 'submitted'
    bill.submitted_on = timezone.localdate()
    bill.save()
    inv = bill.invoice_number or bill.bill_number
    messages.success(request, f'Invoice {inv} marked as submitted (sent to client).')
    return redirect('bill_detail', pk=pk)


@login_required
def bills_submit_bulk(request):
    if request.method != 'POST':
        return redirect('bill_queue_pending')
    profile = get_profile(request.user)
    if not profile.can_submit_bill and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to submit bills.')
        return redirect('bill_queue_pending')
    ids = []
    for x in request.POST.getlist('bill_ids'):
        if str(x).isdigit():
            ids.append(int(x))
    if not ids:
        messages.warning(request, 'No bills selected.')
        return redirect('bill_queue_pending')
    today = timezone.localdate()
    n = Bill.objects.filter(
        pk__in=ids,
        status='pending',
        bill_period_to__isnull=False,
        bill_period_to__lt=today,
    ).update(status='submitted', submitted_on=today)
    messages.success(request, f'{n} bill(s) marked as submitted.')
    return redirect('bill_queue_pending')


@login_required
def bills_mark_paid_bulk(request):
    if request.method != 'POST':
        return redirect('bill_queue_submitted')
    profile = get_profile(request.user)
    if not profile.can_mark_bill_paid and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to mark bills as paid.')
        return redirect('bill_queue_submitted')
    ids = []
    for x in request.POST.getlist('bill_ids'):
        if str(x).isdigit():
            ids.append(int(x))
    if not ids:
        messages.warning(request, 'No bills selected.')
        return redirect('bill_queue_submitted')
    today = date.today()
    n = Bill.objects.filter(pk__in=ids, status='submitted').update(
        status='paid', payment_date=today
    )
    messages.success(request, f'{n} bill(s) marked as paid.')
    return redirect('bill_queue_submitted')


@login_required
def bill_queue_pending(request):
    profile = get_profile(request.user)
    today = timezone.localdate()
    base = Bill.objects.filter(
        status='pending',
        bill_period_to__isnull=False,
        bill_period_to__lt=today,
    )
    return _render_queue_grouped(request, base, 'bills/bill_queue_pending.html', profile)


@login_required
def bill_queue_submitted(request):
    profile = get_profile(request.user)
    base = Bill.objects.filter(status='submitted')
    return _render_queue_grouped(request, base, 'bills/bill_queue_submitted.html', profile)


@login_required
def bill_queue_paid(request):
    profile = get_profile(request.user)
    base = Bill.objects.filter(status='paid')
    return _render_queue_grouped(request, base, 'bills/bill_queue_paid.html', profile)


@login_required
def bill_delete(request, pk):
    profile = get_profile(request.user)
    if not profile.can_delete_bill and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to delete bills.')
        return redirect('bill_detail', pk=pk)
    bill = get_object_or_404(Bill, pk=pk)
    if request.method != 'POST':
        messages.error(request, 'Use the delete button to remove a bill.')
        return redirect('bill_detail', pk=pk)
    inv = bill.invoice_number or bill.bill_number
    pk_str = str(bill.pk)
    if bill.client_id:
        object_repr = f'{inv} — {bill.client.name}'[:500]
    else:
        object_repr = str(inv)[:500]
    ip_addr = request.META.get('HTTP_X_FORWARDED_FOR', '').split(',')[0].strip() or request.META.get('REMOTE_ADDR')
    AuditLog.objects.create(
        user=request.user,
        action=AuditLog.ACTION_DELETE,
        target_model='Bill',
        object_pk=pk_str,
        object_repr=object_repr,
        ip_address=(ip_addr or '')[:45],
    )
    bill.delete()
    messages.success(request, f'Invoice {inv} deleted.')
    return redirect('bill_list')
