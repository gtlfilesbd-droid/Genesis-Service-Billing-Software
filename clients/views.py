from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
import json
from django.http import HttpResponse
from .models import Client, Company, Agreement, AgreementTitlePreset, Service

AGREEMENT_TITLE_CHOICE_OTHER = '__other__'


def _active_agreement_title_presets():
    return AgreementTitlePreset.objects.filter(is_active=True).order_by('sort_order', 'title')


def _resolve_agreement_title_from_post(post):
    choice = (post.get('title_choice') or '').strip()
    custom = (post.get('title_custom') or '').strip()
    if not choice:
        return None, 'Please select an agreement title.'
    if choice == AGREEMENT_TITLE_CHOICE_OTHER:
        if not custom:
            return None, 'Please enter a custom agreement title when "Other" is selected.'
        if len(custom) > 255:
            return None, 'Agreement title must be 255 characters or fewer.'
        return custom, None
    try:
        pk = int(choice)
    except (TypeError, ValueError):
        return None, 'Invalid agreement title selection.'
    preset = AgreementTitlePreset.objects.filter(pk=pk, is_active=True).first()
    if not preset:
        return None, 'Invalid or inactive agreement title.'
    return preset.title, None


def _agreement_title_display_context(agreement=None, post=None):
    presets = _active_agreement_title_presets()
    ctx = {
        'agreement_title_presets': presets,
        'agreement_title_choice_other': AGREEMENT_TITLE_CHOICE_OTHER,
        'selected_title_choice': '',
        'title_is_other': False,
        'posted_title_custom': '',
    }
    if post is not None:
        ctx['selected_title_choice'] = (post.get('title_choice') or '').strip()
        ctx['title_is_other'] = ctx['selected_title_choice'] == AGREEMENT_TITLE_CHOICE_OTHER
        ctx['posted_title_custom'] = post.get('title_custom') or ''
        return ctx
    if agreement and agreement.title:
        match = presets.filter(title=agreement.title).first()
        if match:
            ctx['selected_title_choice'] = str(match.pk)
        else:
            ctx['selected_title_choice'] = AGREEMENT_TITLE_CHOICE_OTHER
            ctx['title_is_other'] = True
            ctx['posted_title_custom'] = agreement.title
    return ctx
from accounts.models import UserProfile, AuditLog


def get_user_profile(user):
    try:
        return user.profile
    except UserProfile.DoesNotExist:
        return UserProfile.objects.create(user=user)


def _agreement_service_groups(agreement):
    groups = []
    for s in agreement.services.all():
        names = [n.strip() for n in (s.name or '').splitlines() if n.strip()]
        groups.append({
            'names': names if names else [''],
            'service_type': s.service_type,
            'charge': s.charge,
            'description': s.description or '',
        })
    return groups


@login_required
def client_list(request):
    clients = Client.objects.all()
    search = request.GET.get('search', '')
    if search:
        clients = clients.filter(name__icontains=search) | clients.filter(short_form__icontains=search)
    paginator = Paginator(clients, 10)
    page = request.GET.get('page')
    clients = paginator.get_page(page)
    profile = get_user_profile(request.user)
    return render(request, 'clients/client_list.html', {'clients': clients, 'search': search, 'profile': profile})


@login_required
def client_add(request):
    profile = get_user_profile(request.user)
    if not profile.can_add_client and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to add clients.')
        return redirect('client_list')

    if request.method == 'POST':
        client = Client(
            name=request.POST.get('name'),
            short_form=request.POST.get('short_form'),
            email=request.POST.get('email'),
            phone=request.POST.get('phone'),
            address=request.POST.get('address'),
            city=request.POST.get('city'),
            country=request.POST.get('country', 'Bangladesh'),
            created_by=request.user
        )
        client.save()
        messages.success(request, f'Client "{client.name}" added successfully.')
        return redirect('client_detail', pk=client.pk)

    return render(request, 'clients/client_form.html', {'action': 'Add', 'profile': profile})


@login_required
def client_detail(request, pk):
    client = get_object_or_404(Client, pk=pk)
    profile = get_user_profile(request.user)
    agreements = client.agreements.select_related('agreement_with').prefetch_related('services').all()
    service_type_choices = dict(Service._meta.get_field('service_type').choices)
    # Each Service row may contain multiple names (one per line).
    from billing.bill_period import count_anniversary_periods, count_monthly_anniversary_periods

    for ag in agreements:
        end_d = ag.effective_amc_end_date()
        months_span = Agreement._months_in_period(ag.start_date, end_d) if ag.start_date else 0
        months_ann = (
            count_monthly_anniversary_periods(ag.start_date, end_d) if ag.start_date else 0
        )
        qtrs = count_anniversary_periods(ag.start_date, end_d, 3) if ag.start_date else 0
        semis = count_anniversary_periods(ag.start_date, end_d, 6) if ag.start_date else 0
        ann = count_anniversary_periods(ag.start_date, end_d, 12) if ag.start_date else 0
        ag.grouped_services = []
        ag.total_monthly_amount = 0
        ag.total_yearly_amount = 0
        for s in ag.services.all():
            from decimal import Decimal, InvalidOperation
            names = [n.strip() for n in (s.name or '').splitlines() if n.strip()]
            try:
                charge = Decimal(str(s.charge or 0))
            except (InvalidOperation, TypeError, ValueError):
                charge = Decimal('0')

            st = (s.service_type or '').lower()
            if st == 'monthly':
                monthly_amount = charge
                yearly_amount = charge * Decimal('12')
            elif st in ('annual', 'yearly'):
                # charge is treated as yearly
                monthly_amount = charge / Decimal('12') if charge else Decimal('0')
                yearly_amount = charge
            elif st == 'quarterly':
                monthly_amount = charge / Decimal('3') if charge else Decimal('0')
                yearly_amount = charge * Decimal('4')
            elif st == 'semi_annual':
                monthly_amount = charge / Decimal('6') if charge else Decimal('0')
                yearly_amount = charge * Decimal('2')
            elif st == 'one_time':
                monthly_amount = charge
                yearly_amount = charge
            else:
                monthly_amount = charge
                yearly_amount = charge * Decimal('12')

            if st in ('annual', 'yearly'):
                total_amc = charge * Decimal(ann or 0)
            elif st == 'monthly':
                total_amc = monthly_amount * Decimal(months_ann or 0)
            elif st == 'quarterly':
                total_amc = charge * Decimal(qtrs or 0)
            elif st == 'semi_annual':
                total_amc = charge * Decimal(semis or 0)
            elif st == 'one_time':
                total_amc = charge
            else:
                total_amc = charge

            ag.total_monthly_amount += monthly_amount
            ag.total_yearly_amount += yearly_amount
            ag.grouped_services.append({
                'names': names if names else ['—'],
                'service_type': s.service_type,
                'service_type_label': service_type_choices.get(s.service_type, s.service_type),
                'charge': charge,
                'monthly_amount': monthly_amount,
                'yearly_amount': yearly_amount,
                'total_amc_amount': total_amc,
                'description': s.description or '',
            })
    return render(request, 'clients/client_detail.html', {
        'client': client,
        'agreements': agreements,
        'profile': profile
    })


@login_required
def agreement_sheet_excel(request, pk):
    """
    Download an Excel sheet for a single agreement's services
    in the same format as the view table (Service/Type/Charge/Monthly/Yearly/Total AMC/Description).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        messages.error(request, 'openpyxl not installed.')
        return redirect('client_list')

    agreement = get_object_or_404(
        Agreement.objects.select_related('client', 'agreement_with').prefetch_related('services'),
        pk=pk,
    )

    from billing.bill_period import count_anniversary_periods, count_monthly_anniversary_periods

    end_d = agreement.effective_amc_end_date()
    months_span = (
        Agreement._months_in_period(agreement.start_date, end_d) if agreement.start_date else 0
    )
    months_ann = (
        count_monthly_anniversary_periods(agreement.start_date, end_d)
        if agreement.start_date
        else 0
    )
    qtrs = count_anniversary_periods(agreement.start_date, end_d, 3) if agreement.start_date else 0
    semis = count_anniversary_periods(agreement.start_date, end_d, 6) if agreement.start_date else 0
    ann = count_anniversary_periods(agreement.start_date, end_d, 12) if agreement.start_date else 0

    service_type_choices = dict(Service._meta.get_field('service_type').choices)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Agreement-{agreement.pk}"

    bf = Font(bold=True)
    header_fill = PatternFill('solid', fgColor='1565C0')
    header_font = Font(bold=True, color='FFFFFF')

    # Title
    ws.merge_cells('A1:G1')
    ws['A1'] = f"{agreement.client.name} — {agreement.title}"
    ws['A1'].font = Font(size=14, bold=True, color='1565C0')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:G2')
    ws['A2'] = f"AMC Period: {agreement.start_date} — {agreement.end_date or 'Ongoing'}"
    ws['A2'].alignment = Alignment(horizontal='center')

    if agreement.agreement_with:
        aw = agreement.agreement_with
        ws.merge_cells('A3:G3')
        sf = (aw.short_form or '').strip()
        ws['A3'] = f"Agreement with: {aw.name}" + (f" ({sf})" if sf else '')
        ws['A3'].alignment = Alignment(horizontal='center')

    ws.append([])

    headers = ['SERVICE NAME(S)', 'TYPE', 'CHARGE (BDT)', 'MONTHLY (BDT)', 'YEARLY (BDT)', 'TOTAL AMC (BDT)', 'DESCRIPTION']
    ws.append(headers)
    hr = ws.max_row
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=hr, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    total_monthly = 0
    total_yearly = 0
    total_amc = 0

    for s in agreement.services.all():
        names = "\n".join([n.strip() for n in (s.name or '').splitlines() if n.strip()]) or '—'
        st = (s.service_type or '').lower()
        charge = float(s.charge or 0)

        # Monthly/Yearly shown amounts (based on type)
        if st in ('annual', 'yearly'):
            monthly_amt = charge / 12 if charge else 0
            yearly_amt = charge
            total_amt = charge * (ann or 0)
        elif st == 'monthly':
            monthly_amt = charge
            yearly_amt = charge * 12
            total_amt = charge * (months_ann or 0)
        elif st == 'quarterly':
            monthly_amt = charge / 3 if charge else 0
            yearly_amt = charge * 4
            total_amt = charge * (qtrs or 0)
        elif st == 'semi_annual':
            monthly_amt = charge / 6 if charge else 0
            yearly_amt = charge * 2
            total_amt = charge * (semis or 0)
        elif st == 'one_time':
            monthly_amt = charge
            yearly_amt = charge
            total_amt = charge
        else:
            monthly_amt = charge
            yearly_amt = charge * 12
            total_amt = charge

        total_monthly += monthly_amt
        total_yearly += yearly_amt
        total_amc += total_amt

        ws.append([
            names,
            service_type_choices.get(s.service_type, s.service_type),
            charge,
            monthly_amt,
            yearly_amt,
            total_amt,
            (s.description or ''),
        ])

    # Totals row
    ws.append([])
    ws.append(['', '', '', 'TOTAL MONTHLY', total_monthly, '', ''])
    ws.append(['', '', '', 'TOTAL YEARLY', total_yearly, '', ''])
    ws.append(['', '', '', 'TOTAL AMC', total_amc, '', ''])
    for r in range(ws.max_row - 2, ws.max_row + 1):
        ws.cell(row=r, column=4).font = bf
        ws.cell(row=r, column=5).font = bf

    # Formatting
    for row in ws.iter_rows(min_row=hr + 1, min_col=1, max_col=7):
        row[0].alignment = Alignment(wrap_text=True, vertical='top')
        row[6].alignment = Alignment(wrap_text=True, vertical='top')
        for c in row[1:6]:
            c.alignment = Alignment(vertical='top')

    for col, w in zip('ABCDEFG', [34, 16, 14, 14, 14, 16, 28]):
        ws.column_dimensions[col].width = w

    filename = f"Agreement-{agreement.pk}-Services.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def client_edit(request, pk):
    profile = get_user_profile(request.user)
    if not profile.can_edit_client and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to edit clients.')
        return redirect('client_detail', pk=pk)

    client = get_object_or_404(Client, pk=pk)
    if request.method == 'POST':
        client.name = request.POST.get('name')
        client.short_form = request.POST.get('short_form')
        client.email = request.POST.get('email')
        client.phone = request.POST.get('phone')
        client.address = request.POST.get('address')
        client.city = request.POST.get('city')
        client.country = request.POST.get('country', 'Bangladesh')
        client.is_active = request.POST.get('is_active') == 'on'
        client.save()
        messages.success(request, 'Client updated successfully.')
        return redirect('client_detail', pk=pk)

    return render(request, 'clients/client_form.html', {'action': 'Edit', 'client': client, 'profile': profile})


@login_required
def client_delete(request, pk):
    profile = get_user_profile(request.user)
    if not profile.can_delete_client and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to delete clients.')
        return redirect('client_detail', pk=pk)

    if request.method != 'POST':
        messages.error(request, 'Invalid request method.')
        return redirect('client_detail', pk=pk)

    client = get_object_or_404(Client, pk=pk)
    name = client.name
    client_pk_str = str(client.pk)
    ip_addr = request.META.get('HTTP_X_FORWARDED_FOR', '').split(',')[0].strip() or request.META.get('REMOTE_ADDR')
    AuditLog.objects.create(
        user=request.user,
        action=AuditLog.ACTION_DELETE,
        target_model='Client',
        object_pk=client_pk_str,
        object_repr=name[:500],
        ip_address=(ip_addr or '')[:45],
    )
    client.delete()
    messages.success(request, f'Client "{name}" deleted successfully.')
    return redirect('client_list')


@login_required
def agreement_add(request, client_pk):
    profile = get_user_profile(request.user)
    if not profile.can_add_client and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to add agreements.')
        return redirect('client_detail', pk=client_pk)

    client = get_object_or_404(Client, pk=client_pk)
    companies = Company.objects.filter(is_active=True).order_by('name')

    if request.method == 'POST':
        aw_id = request.POST.get('agreement_with')
        if not aw_id:
            messages.error(request, 'Please select Agreement With.')
            ctx = {
                'client': client,
                'profile': profile,
                'service_types': Service._meta.get_field('service_type').choices,
                'companies': companies,
                **_agreement_title_display_context(post=request.POST),
            }
            return render(request, 'clients/agreement_form.html', ctx)
        agreement_with = Company.objects.filter(pk=aw_id, is_active=True).first()
        if not agreement_with:
            messages.error(request, 'Invalid company selected for Agreement With.')
            ctx = {
                'client': client,
                'profile': profile,
                'service_types': Service._meta.get_field('service_type').choices,
                'companies': companies,
                **_agreement_title_display_context(post=request.POST),
            }
            return render(request, 'clients/agreement_form.html', ctx)

        resolved_title, title_err = _resolve_agreement_title_from_post(request.POST)
        if title_err:
            messages.error(request, title_err)
            ctx = {
                'client': client,
                'profile': profile,
                'service_types': Service._meta.get_field('service_type').choices,
                'companies': companies,
                **_agreement_title_display_context(post=request.POST),
            }
            return render(request, 'clients/agreement_form.html', ctx)

        if not (request.POST.get('end_date') or '').strip():
            messages.error(request, 'Please select End Date.')
            ctx = {
                'client': client,
                'profile': profile,
                'service_types': Service._meta.get_field('service_type').choices,
                'companies': companies,
                **_agreement_title_display_context(post=request.POST),
            }
            return render(request, 'clients/agreement_form.html', ctx)

        agreement = Agreement(
            client=client,
            agreement_with=agreement_with,
            title=resolved_title,
            start_date=request.POST.get('start_date'),
            end_date=request.POST.get('end_date') or None,
            notes=request.POST.get('notes'),
            attachment=request.FILES.get('attachment'),
            vat_ait_excluded=request.POST.get('vat_ait_excluded') == 'on',
            created_by=request.user
        )
        agreement.save()

        # Add services
        service_names_json = request.POST.getlist('service_names_json[]')
        service_types = request.POST.getlist('service_type[]')
        service_charges = request.POST.getlist('service_charge[]')
        service_descs = request.POST.getlist('service_description[]')

        # Backward compatibility: if older template posts flat service_name[]
        if not service_names_json:
            service_names = request.POST.getlist('service_name[]')
            for i in range(len(service_names)):
                if service_names[i]:
                    Service.objects.create(
                        agreement=agreement,
                        name=service_names[i],
                        service_type=service_types[i] if i < len(service_types) else 'monthly',
                        charge=service_charges[i] if i < len(service_charges) else 0,
                        description=service_descs[i] if i < len(service_descs) else '',
                    )
        else:
            for i in range(len(service_names_json)):
                raw = service_names_json[i]
                try:
                    names = json.loads(raw) if raw else []
                except json.JSONDecodeError:
                    names = []
                cleaned = [n.strip() for n in names if isinstance(n, str) and n.strip()]
                if not cleaned:
                    continue
                # One row = one charge. Store multiple names in one Service (one per line).
                Service.objects.create(
                    agreement=agreement,
                    name="\n".join(cleaned),
                    service_type=service_types[i] if i < len(service_types) else 'monthly',
                    charge=service_charges[i] if i < len(service_charges) else 0,
                    description=service_descs[i] if i < len(service_descs) else '',
                )

        messages.success(request, 'Agreement added successfully.')
        return redirect('client_detail', pk=client_pk)

    return render(request, 'clients/agreement_form.html', {
        'client': client,
        'profile': profile,
        'service_types': Service._meta.get_field('service_type').choices,
        'companies': companies,
        **_agreement_title_display_context(),
    })


@login_required
def agreement_edit(request, pk):
    profile = get_user_profile(request.user)
    agreement = get_object_or_404(Agreement, pk=pk)
    client = agreement.client
    companies = Company.objects.filter(is_active=True).order_by('name')

    if not profile.can_edit_client and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to edit agreements.')
        return redirect('client_detail', pk=client.pk)

    if request.method == 'POST':
        aw_id = request.POST.get('agreement_with')
        if not aw_id:
            messages.error(request, 'Please select Agreement With.')
            ctx = {
                'client': client,
                'agreement': agreement,
                'profile': profile,
                'service_types': Service._meta.get_field('service_type').choices,
                'service_groups': _agreement_service_groups(agreement),
                'companies': companies,
                **_agreement_title_display_context(post=request.POST),
            }
            return render(request, 'clients/agreement_form.html', ctx)
        agreement_with = Company.objects.filter(pk=aw_id, is_active=True).first()
        if not agreement_with:
            messages.error(request, 'Invalid company selected for Agreement With.')
            ctx = {
                'client': client,
                'agreement': agreement,
                'profile': profile,
                'service_types': Service._meta.get_field('service_type').choices,
                'service_groups': _agreement_service_groups(agreement),
                'companies': companies,
                **_agreement_title_display_context(post=request.POST),
            }
            return render(request, 'clients/agreement_form.html', ctx)

        resolved_title, title_err = _resolve_agreement_title_from_post(request.POST)
        if title_err:
            messages.error(request, title_err)
            ctx = {
                'client': client,
                'agreement': agreement,
                'profile': profile,
                'service_types': Service._meta.get_field('service_type').choices,
                'service_groups': _agreement_service_groups(agreement),
                'companies': companies,
                **_agreement_title_display_context(post=request.POST),
            }
            return render(request, 'clients/agreement_form.html', ctx)

        if not (request.POST.get('end_date') or '').strip():
            messages.error(request, 'Please select End Date.')
            ctx = {
                'client': client,
                'agreement': agreement,
                'profile': profile,
                'service_types': Service._meta.get_field('service_type').choices,
                'service_groups': _agreement_service_groups(agreement),
                'companies': companies,
                **_agreement_title_display_context(post=request.POST),
            }
            return render(request, 'clients/agreement_form.html', ctx)

        agreement.agreement_with = agreement_with
        agreement.title = resolved_title
        agreement.start_date = request.POST.get('start_date')
        agreement.end_date = request.POST.get('end_date') or None
        agreement.notes = request.POST.get('notes')
        agreement.is_active = request.POST.get('is_active') == 'on'
        agreement.vat_ait_excluded = request.POST.get('vat_ait_excluded') == 'on'
        if request.FILES.get('attachment'):
            agreement.attachment = request.FILES.get('attachment')
        agreement.save()

        # Update services
        agreement.services.all().delete()
        service_names_json = request.POST.getlist('service_names_json[]')
        service_types = request.POST.getlist('service_type[]')
        service_charges = request.POST.getlist('service_charge[]')
        service_descs = request.POST.getlist('service_description[]')

        # Backward compatibility: if older template posts flat service_name[]
        if not service_names_json:
            service_names = request.POST.getlist('service_name[]')
            for i in range(len(service_names)):
                if service_names[i]:
                    Service.objects.create(
                        agreement=agreement,
                        name=service_names[i],
                        service_type=service_types[i] if i < len(service_types) else 'monthly',
                        charge=service_charges[i] if i < len(service_charges) else 0,
                        description=service_descs[i] if i < len(service_descs) else '',
                    )
        else:
            for i in range(len(service_names_json)):
                raw = service_names_json[i]
                try:
                    names = json.loads(raw) if raw else []
                except json.JSONDecodeError:
                    names = []
                cleaned = [n.strip() for n in names if isinstance(n, str) and n.strip()]
                if not cleaned:
                    continue
                Service.objects.create(
                    agreement=agreement,
                    name="\n".join(cleaned),
                    service_type=service_types[i] if i < len(service_types) else 'monthly',
                    charge=service_charges[i] if i < len(service_charges) else 0,
                    description=service_descs[i] if i < len(service_descs) else '',
                )

        messages.success(request, 'Agreement updated successfully.')
        return redirect('client_detail', pk=client.pk)

    return render(request, 'clients/agreement_form.html', {
        'client': client,
        'agreement': agreement,
        'profile': profile,
        'service_types': Service._meta.get_field('service_type').choices,
        'service_groups': _agreement_service_groups(agreement),
        'companies': companies,
        **_agreement_title_display_context(agreement=agreement),
    })


@login_required
def agreement_delete(request, pk):
    profile = get_user_profile(request.user)
    agreement = get_object_or_404(Agreement, pk=pk)
    client = agreement.client

    if not profile.can_edit_client and not request.user.is_superuser:
        messages.error(request, 'You do not have permission to delete agreements.')
        return redirect('client_detail', pk=client.pk)

    if request.method != 'POST':
        messages.error(request, 'Invalid request method.')
        return redirect('client_detail', pk=client.pk)

    title = agreement.title
    agreement.delete()
    messages.success(request, f'Agreement "{title}" deleted successfully.')
    return redirect('client_detail', pk=client.pk)
